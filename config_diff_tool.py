#!/usr/bin/env python3
"""
Configuration Diff Tool

A tool to recursively compare configuration files across multiple server directories
and report differences in an Excel format. Supports nested directory structures.

Directory Structure Example:
    APP/
      server1/
        profiles/site.xml
        rc/mongo.rc
      server2/
        profiles/site.xml
        rc/mongo.rc

Usage:
    python config_diff_tool.py <directory_path> [--output output.xlsx] [--verbose]
"""

import os
import sys
import argparse
import logging
from pathlib import Path
from collections import defaultdict, OrderedDict
from typing import Dict, List, Set, Tuple, Any
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.utils.dataframe import dataframe_to_rows


class ConfigDiffTool:
    """Main class for comparing configuration files across server directories."""
    
    def __init__(self, base_directory: str, output_file: str = "config_diff_report.xlsx"):
        self.base_directory = Path(base_directory)
        self.output_file = output_file
        self.config_extensions = {'.rc', '.xml', '.jrc'}
        self.host_configs = defaultdict(dict)  # {host: {filename: OrderedDict{key: value}}}
        self.all_files = set()
        self.all_keys_per_file = defaultdict(list)  # Changed to list to preserve order
        self.file_key_order = defaultdict(list)  # Track the order keys appear in each file
        
        # Set up logging
        logging.basicConfig(
            level=logging.INFO,
            format='%(asctime)s - %(levelname)s - %(message)s'
        )
        self.logger = logging.getLogger(__name__)
    
    def is_valid_config_file(self, file_path: Path) -> bool:
        """Check if a file is a valid configuration file based on extension."""
        return file_path.suffix.lower() in self.config_extensions
    
    def parse_config_file(self, file_path: Path) -> OrderedDict[str, str]:
        """
        Parse a configuration file into key-value pairs, preserving order.
        
        Args:
            file_path: Path to the configuration file
            
        Returns:
            OrderedDict of key-value pairs in the order they appear in the file
        """
        config_data = OrderedDict()
        
        try:
            with open(file_path, 'r', encoding='utf-8', errors='ignore') as file:
                for line_num, line in enumerate(file, 1):
                    # Strip whitespace
                    line = line.strip()
                    
                    # Skip empty lines and comments
                    if not line or line.startswith('#'):
                        continue
                    
                    # Split by '=' and ensure we have both key and value
                    if '=' in line:
                        key, value = line.split('=', 1)  # Split only on first '='
                        key = key.strip()
                        value = value.strip()
                        
                        if key:  # Only add if key is not empty
                            config_data[key] = value
                            
        except Exception as e:
            self.logger.warning(f"Error parsing {file_path}: {e}")
            
        return config_data
    
    def scan_directories(self) -> None:
        """Recursively scan the base directory for host subdirectories and their config files."""
        if not self.base_directory.exists():
            raise FileNotFoundError(f"Directory {self.base_directory} does not exist")
        
        self.logger.info(f"Recursively scanning directory: {self.base_directory}")
        
        # Find all subdirectories (host directories) at the first level
        host_directories = [d for d in self.base_directory.iterdir() 
                          if d.is_dir() and not d.name.startswith('.')]
        
        if not host_directories:
            raise ValueError("No host directories found in the specified path")
        
        self.logger.info(f"Found {len(host_directories)} host directories")
        
        # Process each host directory recursively
        for host_dir in host_directories:
            host_name = host_dir.name
            self.logger.info(f"Processing host: {host_name}")
            
            # Recursively find all config files in this host directory and subdirectories
            config_files_found = 0
            for config_file in host_dir.rglob('*'):
                if config_file.is_file() and self.is_valid_config_file(config_file):
                    # Get the relative path from the host directory to maintain file identity
                    relative_path = config_file.relative_to(host_dir)
                    
                    # Use the full relative path as file identifier to handle files with same name in different subdirs
                    file_identifier = str(relative_path).replace('\\', '/')  # Normalize path separators
                    
                    self.all_files.add(file_identifier)
                    config_files_found += 1
                    
                    # Parse the configuration file
                    config_data = self.parse_config_file(config_file)
                    self.host_configs[host_name][file_identifier] = config_data
                    
                    # Track key order for this file (use the first host that has this file)
                    if file_identifier not in self.file_key_order:
                        self.file_key_order[file_identifier] = list(config_data.keys())
                    else:
                        # Add any new keys that weren't in the first file we saw
                        existing_keys = set(self.file_key_order[file_identifier])
                        for key in config_data.keys():
                            if key not in existing_keys:
                                self.file_key_order[file_identifier].append(key)
                    
                    # Update all keys for this file (maintaining order)
                    for key in config_data.keys():
                        if key not in self.all_keys_per_file[file_identifier]:
                            self.all_keys_per_file[file_identifier].append(key)
                    
                    self.logger.debug(f"Parsed {file_identifier} for {host_name}: {len(config_data)} keys")
            
            self.logger.info(f"Found {config_files_found} config files in {host_name}")
    
    def find_differences(self) -> List[Dict[str, Any]]:
        """
        Find differences in configuration values across hosts.
        
        Returns:
            List of all differences with file name included
        """
        all_differences = []
        
        for file_name in sorted(self.all_files):
            # Use the preserved order instead of sorting
            all_keys = self.file_key_order[file_name]
            
            for key in all_keys:
                # Collect values for this key across all hosts
                key_values = {}
                hosts_with_key = []
                
                for host_name in sorted(self.host_configs.keys()):
                    if file_name in self.host_configs[host_name]:
                        if key in self.host_configs[host_name][file_name]:
                            value = self.host_configs[host_name][file_name][key]
                            key_values[host_name] = value
                            hosts_with_key.append(host_name)
                        else:
                            key_values[host_name] = "** MISSING **"
                    else:
                        key_values[host_name] = "** FILE NOT FOUND **"
                
                # Check if there are differences in values
                unique_values = set(v for v in key_values.values() 
                                  if v not in ["** MISSING **", "** FILE NOT FOUND **"])
                
                if len(unique_values) > 1 or len(key_values) != len(hosts_with_key):
                    # There are differences
                    diff_entry = {
                        'file_name': file_name,
                        'key': key,
                        'hosts': key_values,
                        'unique_values': list(unique_values),
                        'has_missing': "** MISSING **" in key_values.values(),
                        'has_missing_file': "** FILE NOT FOUND **" in key_values.values()
                    }
                    all_differences.append(diff_entry)
                
        return all_differences
    
    def create_excel_report(self, differences: List[Dict[str, Any]]) -> None:
        """Create an Excel report with all differences on one sheet."""
        wb = Workbook()
        
        # Remove default worksheet
        wb.remove(wb.active)
        
        # Create summary worksheet
        summary_ws = wb.create_sheet("Summary")
        self._create_summary_sheet(summary_ws, differences)
        
        # Create single consolidated differences worksheet
        if differences:
            diff_ws = wb.create_sheet("All Differences")
            self._create_consolidated_diff_sheet(diff_ws, differences)
        
        # Create host overview worksheet
        overview_ws = wb.create_sheet("Host Overview")
        self._create_host_overview_sheet(overview_ws)
        
        # Save the workbook
        wb.save(self.output_file)
        self.logger.info(f"Excel report saved to: {self.output_file}")
    
    def _create_summary_sheet(self, ws, differences: List[Dict[str, Any]]) -> None:
        """Create the summary worksheet."""
        ws.title = "Summary"
        
        # Header
        ws['A1'] = "Configuration Diff Tool - Summary Report"
        ws['A1'].font = Font(bold=True, size=14)
        
        # Statistics
        row = 3
        ws[f'A{row}'] = "Statistics:"
        ws[f'A{row}'].font = Font(bold=True)
        row += 1
        
        ws[f'A{row}'] = f"Total hosts analyzed: {len(self.host_configs)}"
        row += 1
        ws[f'A{row}'] = f"Total config files: {len(self.all_files)}"
        row += 1
        
        # Count files with differences
        files_with_diffs = set(diff['file_name'] for diff in differences)
        ws[f'A{row}'] = f"Files with differences: {len(files_with_diffs)}"
        row += 1
        ws[f'A{row}'] = f"Total differences found: {len(differences)}"
        row += 2
        
        # Files with differences
        if files_with_diffs:
            ws[f'A{row}'] = "Files with differences:"
            ws[f'A{row}'].font = Font(bold=True)
            row += 1
            
            ws[f'A{row}'] = "File Name"
            ws[f'B{row}'] = "Keys with Differences"
            ws[f'A{row}'].font = Font(bold=True)
            ws[f'B{row}'].font = Font(bold=True)
            row += 1
            
            # Count differences per file
            file_diff_counts = {}
            for diff in differences:
                file_name = diff['file_name']
                file_diff_counts[file_name] = file_diff_counts.get(file_name, 0) + 1
            
            for file_name in sorted(file_diff_counts.keys()):
                ws[f'A{row}'] = file_name
                ws[f'B{row}'] = file_diff_counts[file_name]
                row += 1
    
    def _create_consolidated_diff_sheet(self, ws, differences: List[Dict[str, Any]]) -> None:
        """Create a single worksheet with all differences."""
        ws.title = "All Differences"
        
        # Header
        ws['A1'] = "All Configuration Differences"
        ws['A1'].font = Font(bold=True, size=12)
        
        # Column headers
        row = 3
        ws[f'A{row}'] = "File Name"
        ws[f'B{row}'] = "Key"
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'B{row}'].font = Font(bold=True)
        
        col = 3
        host_names = sorted(self.host_configs.keys())
        for host_name in host_names:
            ws.cell(row=row, column=col, value=host_name)
            ws.cell(row=row, column=col).font = Font(bold=True)
            col += 1
        
        # Data rows
        row += 1
        for diff in differences:
            ws[f'A{row}'] = diff['file_name']
            ws[f'B{row}'] = diff['key']
            
            # Determine which values are actually different
            # Get all non-error values for this key
            actual_values = [v for v in diff['hosts'].values() 
                           if v not in ["** MISSING **", "** FILE NOT FOUND **"]]
            
            # If there's more than one unique actual value, determine which cells to highlight
            values_to_highlight = set()
            if len(set(actual_values)) > 1:
                # Find the most common value (if any)
                value_counts = {}
                for val in actual_values:
                    value_counts[val] = value_counts.get(val, 0) + 1
                
                # If there's a clear majority value, highlight only the minority values
                # Otherwise, highlight all values that differ from each other
                max_count = max(value_counts.values()) if value_counts else 0
                majority_values = [val for val, count in value_counts.items() if count == max_count]
                
                if len(majority_values) == 1 and max_count > 1:
                    # There's a clear majority value, highlight only the different ones
                    majority_value = majority_values[0]
                    values_to_highlight = set(val for val in actual_values if val != majority_value)
                else:
                    # No clear majority, highlight all different values
                    values_to_highlight = set(actual_values)
            
            col = 3
            for host_name in host_names:
                value = diff['hosts'].get(host_name, "** NOT FOUND **")
                cell = ws.cell(row=row, column=col, value=value)
                
                # Color coding - more precise highlighting
                if value == "** MISSING **":
                    cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")  # Yellow
                elif value == "** FILE NOT FOUND **":
                    cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")  # Red
                elif value in values_to_highlight:
                    cell.fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")  # Orange
                
                col += 1
            row += 1
    
    def _create_host_overview_sheet(self, ws) -> None:
        """Create a host overview worksheet."""
        ws.title = "Host Overview"
        
        # Header
        ws['A1'] = "Host Overview"
        ws['A1'].font = Font(bold=True, size=12)
        
        # Column headers
        row = 3
        ws[f'A{row}'] = "Host Name"
        ws[f'B{row}'] = "Config Files Found"
        ws[f'C{row}'] = "Total Keys"
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'B{row}'].font = Font(bold=True)
        ws[f'C{row}'].font = Font(bold=True)
        
        row += 1
        for host_name in sorted(self.host_configs.keys()):
            ws[f'A{row}'] = host_name
            ws[f'B{row}'] = len(self.host_configs[host_name])
            
            # Count total keys across all files for this host
            total_keys = sum(len(file_config) for file_config in self.host_configs[host_name].values())
            ws[f'C{row}'] = total_keys
            row += 1
    
    def run(self) -> None:
        """Main execution method."""
        try:
            self.logger.info("Starting configuration diff analysis...")
            
            # Scan directories and parse files
            self.scan_directories()
            
            # Find differences
            self.logger.info("Analyzing differences...")
            differences = self.find_differences()
            
            if not differences:
                self.logger.info("No differences found across all configuration files!")
                # Still create a report showing this
                wb = Workbook()
                ws = wb.active
                ws.title = "No Differences Found"
                ws['A1'] = "No configuration differences found across all hosts!"
                ws['A1'].font = Font(bold=True, size=14)
                wb.save(self.output_file)
            else:
                # Create Excel report
                self.logger.info(f"Found {len(differences)} total differences")
                self.create_excel_report(differences)
            
            self.logger.info("Analysis complete!")
            
        except Exception as e:
            self.logger.error(f"Error during analysis: {e}")
            raise


def main():
    """Main function to run the configuration diff tool."""
    parser = argparse.ArgumentParser(
        description="Compare configuration files across server directories",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  python config_diff_tool.py /path/to/servers
  python config_diff_tool.py /path/to/servers --output my_report.xlsx
  python config_diff_tool.py /path/to/servers --verbose
        """
    )
    
    parser.add_argument(
        'directory',
        help='Path to directory containing server subdirectories'
    )
    
    parser.add_argument(
        '--output', '-o',
        default='config_diff_report.xlsx',
        help='Output Excel file name (default: config_diff_report.xlsx)'
    )
    
    parser.add_argument(
        '--verbose', '-v',
        action='store_true',
        help='Enable verbose logging'
    )
    
    args = parser.parse_args()
    
    # Set logging level
    if args.verbose:
        logging.getLogger().setLevel(logging.DEBUG)
    
    # Validate input directory
    if not os.path.exists(args.directory):
        print(f"Error: Directory '{args.directory}' does not exist")
        sys.exit(1)
    
    # Run the tool
    try:
        tool = ConfigDiffTool(args.directory, args.output)
        tool.run()
        print(f"\nReport generated successfully: {args.output}")
        
    except Exception as e:
        print(f"Error: {e}")
        sys.exit(1)


if __name__ == "__main__":
    main()